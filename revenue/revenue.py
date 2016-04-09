# Monthly revenue sheet
# 上市 sii 上櫃 otc 興櫃 rotc
#http://mops.twse.com.tw/nas/t21/sii/t21sc03_105_2_0.html
#http://mops.twse.com.tw/nas/t21/otc/t21sc03_105_2_0.html
#http://mops.twse.com.tw/nas/t21/rotc/t21sc03_105_2_0.html

from urllib.request import Request, urlopen
from urllib.error import HTTPError
from urllib.parse import urlencode
from bs4 import BeautifulSoup
from shutil import copyfile
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import re
import os

def balance_sheet_link_in_excel(co_id):
	return "=HYPERLINK(\"http://mops.twse.com.tw/mops/web/ajax_t164sb03?encodeURIComponent=1&step=1&firstin=1&off=1&queryName=co_id&TYPEK=all&isnew=true&co_id=" + co_id + "\", \"連結\")"

revenue_sheet_name = 'Revenue'
market_name = {'sii': '上市', 'otc': '上櫃', 'rotc': '興櫃'}
market_header = '市場'
download_time_header = '下載時間'
balance_sheet_header = '合併資產負債表'

current_datetime = datetime.datetime.now().strftime("%Y-%m-%d-%02H-%02M-%02S")

current_year = datetime.datetime.now().year
query_ROC_year = current_year - 1911
current_month = datetime.datetime.now().month

if current_month == 1:
	query_month = 12
	query_ROC_year -= 1
else:
	query_month = current_month - 1

# company_id: income_data_dict
company_income = {}
# header: count
all_headers = ['公司代號', '公司名稱', '當月營收', '上月營收', '去年當月營收', '上月比較增減(%)', '去年同月增減(%)', '當月累計營收', '去年累計營收', '前期比較增減(%)', '備註']

out_dir = 'out/'
if not os.path.exists(out_dir):
	os.makedirs(out_dir)
revenue_file = 'revenue_' + str(query_ROC_year) + '_' + str(query_month) + '.xlsx'
new_revenue_file = out_dir + current_datetime + '_' + str(query_ROC_year) + '_' + str(query_month) + '.xlsx'


for typek in market_name.keys():
	url = 'http://mops.twse.com.tw/nas/t21/' + typek + '/t21sc03_' + str(query_ROC_year) + '_' + str(query_month) + '_0.html'
	save_file = out_dir + current_datetime + '_' + str(query_ROC_year) + '_' + str(query_month) + '_' + typek + '.html'
	print(url)
	response = urlopen(url)
	html = response.read()
	out = open(save_file, "w")
	out.write(html.decode('big5hkscs'))
	out.close()
	soup = BeautifulSoup(html.decode('big5hkscs'), "html.parser")
	[empty_table, global_table] = soup.find_all('table', limit=2)
	tables = global_table.find_all('table')
	for table in tables:
		table = table.find('table')
		#print (table)
		if not table:
			continue
		rows = table.find_all('tr')
		for row in rows:
			cols = row.find_all('td')
			cols = [ele.text.strip() for ele in cols]
			if cols:
				#print (cols)
				income = {market_header: market_name[typek]}
				index = 0
				company_id = cols[0]
				if (re.match("\d{4}", str(company_id))):
					print("id\t" + str(company_id))
					for item in cols:
						income[all_headers[index]] = item
						print(str(all_headers[index]) + "\t" + str(item))
						index += 1
					company_income[company_id] = income
					#print(income)

# Write all headers to excel file
wb = Workbook()
ws = wb.active
ws.title = revenue_sheet_name

x = 1
y = 1
for v in [download_time_header, market_header, balance_sheet_header]:
	ws.cell(row = x, column = y).value = v
	y += 1

for header in all_headers:
	#print(header)
	ws.cell(row = x, column = y).value = header
	#print(str(x)+","+str(y)+","+str(header))
	y += 1

for company in company_income.keys():
	#print("company\t" + str(company))
	income = company_income[company]
	#print(income)
	x += 1
	y = 1
	for v in [str(current_datetime), income.get(market_header, 'n/a'), balance_sheet_link_in_excel(company)]:
		ws.cell(row = x, column = y).value = v
		y += 1
	for header in all_headers:
		value = income.get(header, 'n/a')
		ws.cell(row = x, column = y).value = value
		#print(str(x)+","+str(y)+","+str(value))
		y = y+1

wb.save(new_revenue_file)

if not os.path.exists(revenue_file):
	copyfile(new_revenue_file, revenue_file)
	quit()

#
# Combine with old data
#
company_index_in_columns = 3

wb = load_workbook(filename = revenue_file)
ws = wb.get_sheet_by_name(revenue_sheet_name)
old_companies = []
for col in ws.columns[company_index_in_columns][1:]:
	old_companies.append(col.value)
x = len(old_companies) + 1
#print (old_companies)

for company in company_income.keys():
	if company not in old_companies:
		print (str(company) + "-> new")
		income = company_income[company]
		#print(income)
		x += 1
		y = 1
		for v in [str(current_datetime), income.get(market_header, 'n/a'), balance_sheet_link_in_excel(company)]:
			ws.cell(row = x, column = y).value = v
			y += 1
		for header in all_headers:
			value = income.get(header, 'n/a')
			ws.cell(row = x, column = y).value = value
			#print(str(x)+","+str(y)+","+str(value))
			y = y+1
wb.save(revenue_file)

