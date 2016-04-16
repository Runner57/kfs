# income sheet
# 上市 sii 上櫃 otc 興櫃 rotc

from urllib.request import Request, urlopen
from urllib.error import HTTPError
from urllib.parse import urlencode
from bs4 import BeautifulSoup
from shutil import copyfile
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os

def balance_sheet_link_in_excel(co_id):
	return "=HYPERLINK(\"http://mops.twse.com.tw/mops/web/ajax_t164sb03?encodeURIComponent=1&step=1&firstin=1&off=1&queryName=co_id&TYPEK=all&isnew=true&co_id=" + co_id + "\", \"連結\")"

income_sheet_name = 'Income_Sheet'

url = 'http://mops.twse.com.tw/mops/web/ajax_t163sb04?%s'
encodeURIComponent = '1'
step = '1'
firstin = '1'
eps_header = '基本每股盈餘（元）'
market_header = '市場'
download_time_header = '下載時間'
balance_sheet_header = '合併資產負債表'
market_name = {'sii': '上市', 'otc': '上櫃', 'rotc': '興櫃'}
current_datetime = datetime.datetime.now().strftime("%Y-%m-%d-%02H-%02M-%02S")
current_year = datetime.datetime.now().year
query_ROC_year = '105'
query_quarter = '01'

out_dir = 'out/'
if not os.path.exists(out_dir):
	os.makedirs(out_dir)
income_file = 'income_' + str(query_ROC_year) + '_' + str(query_quarter) + '.xlsx'
new_income_file = out_dir + current_datetime + '_' + str(query_ROC_year) + '_' + str(query_quarter) + '.xlsx'

# company_id: income_data_dict
company_income = {}
# header: count
all_headers = []

for typek in ['sii', 'otc', 'rotc']:
	params = urlencode({'encodeURIComponent': encodeURIComponent, 'step': step, 'firstin': firstin, 'TYPEK': typek, 'year': query_ROC_year, 'season': query_quarter})
	save_file = out_dir + current_datetime + '_' + str(query_ROC_year) + '_' + str(query_quarter) + '_' + typek + '.html'
	response = urlopen(url % params)
	html = response.read()
	out = open(save_file, "w")
	out.write(html.decode('utf-8'))
	out.close()
	soup = BeautifulSoup(html, "html.parser")
	tables = soup.find_all("table")
	for table in tables:
		headers = table.find_all('th')
		headers = [ele.text.strip() for ele in headers]
		if not headers:
			# Skip the table which has no headers
			continue
		for header in headers:
			if not header in all_headers:
				all_headers.append(header)

		rows = table.find_all('tr')
		for row in rows:
		    cols = row.find_all('td')
		    cols = [ele.text.strip() for ele in cols]
		    if cols:
			    #print (cols)
			    income = {market_header: market_name[typek]}
			    index = 0
			    company_id = cols[0]
			    #print("id\t" + str(company_id))
			    for item in cols:
				#    if item == '--':
				#	    item = 'Empty'
				    income[headers[index]] = item
				    #print(str(headers[index]) + "\t" + str(item))
				    index += 1
			    company_income[company_id] = income
			    print(income)

# Write all headers to excel file
wb = Workbook()
ws = wb.active
ws.title = income_sheet_name

x = 1
y = 1
for v in [download_time_header, market_header, balance_sheet_header, eps_header]:
	ws.cell(row = x, column = y).value = v
	y += 1

for header in all_headers:
	#print(header + "\t" + str(headers_count[header]))
	ws.cell(row = x, column = y).value = header
	#print(str(x)+","+str(y)+","+str(header))
	y += 1

for company in company_income.keys():
	#print("company\t" + str(company))
	income = company_income[company]
	#print(income)
	y = 1
	x += 1
	for v in [str(current_datetime), income.get(market_header, 'n/a'), balance_sheet_link_in_excel(company), income.get(eps_header, 'n/a')]:
		ws.cell(row = x, column = y).value = v
		y += 1
	for header in all_headers:
		value = income.get(header, 'n/a')
		ws.cell(row = x, column = y).value = value
		#print(str(x)+","+str(y)+","+str(value))
		y = y+1

wb.save(new_income_file)

if not os.path.exists(income_file):
	copyfile(new_income_file, income_file)
	quit()

#
# Combine with old data
#
company_index_in_columns = 4

wb = load_workbook(filename = income_file)
ws = wb.get_sheet_by_name(income_sheet_name)
old_companies = []
for col in ws.columns[company_index_in_columns][1:]:
	old_companies.append(col.value)
x = len(old_companies) + 1
#print (old_companies)

for company in company_income.keys():
	if company not in old_companies:
		print (str(company) + "-> new")
		income = company_income[company]
		#print(income)
		x += 1
		y = 1
		for v in [str(current_datetime), income.get(market_header, 'n/a'), balance_sheet_link_in_excel(company), income.get(eps_header, 'n/a')]:
			ws.cell(row = x, column = y).value = v
			y += 1
		for header in all_headers:
			value = income.get(header, 'n/a')
			ws.cell(row = x, column = y).value = value
			#print(str(x)+","+str(y)+","+str(value))
			y = y+1
wb.save(income_file)

